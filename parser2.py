from datetime import datetime, timedelta
import re
import openpyxl
from rich.console import Console
from ics import Calendar, Event

console = Console()


def get_cell_value(cell):
    if isinstance(cell.value, str):
        return cell.value.strip()
    elif isinstance(cell.value, (int, float)):
        return str(cell.value)
    elif isinstance(cell.value, bool):
        return str(cell.value)
    elif isinstance(cell.value, datetime):
        return cell.value.strftime("%Y-%m-%d %H:%M:%S")
    elif cell.value is None:
        return ""
    else:
        return "Unknown Value"


def is_merged(sheet, row, col):
    for merged_range in sheet.merged_cells.ranges:
        if sheet.cell(row=row, column=col).coordinate in merged_range:
            return True
    return False


def get_week_class(sheet, row, col, week_type):
    top_value = sheet.cell(row=row, column=col).value
    bottom_value = sheet.cell(row=row + 1, column=col).value

    if is_merged(sheet, row, col):
        return top_value

    if week_type == "чисельник":
        return top_value if top_value else None
    else:
        return bottom_value if bottom_value else None


def extract_groups(sheet):
    group_names = []

    if sheet.cell(row=6, column=2).value == "Пари":
        for col in range(3, sheet.max_column):
            group_names.append(get_cell_value(sheet.cell(row=6, column=col)))
    return group_names


def extract_days(sheet):
    days = []
    skip_phrases = ["В.о. декана", "ДВВС", "І. І. Дияк"]

    for row in range(7, sheet.max_row):
        day = get_cell_value(sheet.cell(row=row, column=1))
        if any(phrase in day for phrase in skip_phrases):
            continue
        if day:
            days.append(day)
    return days


def extract_class_times(sheet):
    class_times = []
    for row in range(7, sheet.max_row + 1):
        time = get_cell_value(sheet.cell(row=row, column=2))
        if time:
            class_times.append(time)
    return class_times


def extract_class_times_by_day(sheet, day):
    class_times = []
    skip_phrases = ["В.о. декана", "ДВВС", "І. І. Дияк"]

    current_day = None
    for row in range(7, sheet.max_row + 1):
        cell_day = get_cell_value(sheet.cell(row=row, column=1))

        if any(phrase in cell_day for phrase in skip_phrases):
            continue

        if cell_day:
            current_day = cell_day

        if current_day == day:
            time = get_cell_value(sheet.cell(row=row, column=2))
            if time:
                class_times.append(time)

    return class_times


def extract_classes_by_group(sheet, group):
    classes = []
    skip_phrases = ["В.о. декана", "ДВВС", "І. І. Дияк"]

    current_day = None
    last_known_time = None

    for row in range(7, sheet.max_row + 1):
        cell_day = get_cell_value(sheet.cell(row=row, column=1))
        if any(phrase in cell_day for phrase in skip_phrases):
            continue

        if cell_day:
            current_day = cell_day

        class_time = get_cell_value(sheet.cell(row=row, column=2))
        if class_time:
            last_known_time = class_time

        for col in range(3, sheet.max_column + 1):
            cell_group = get_cell_value(sheet.cell(row=6, column=col))
            if cell_group == group:
                class_name = get_cell_value(sheet.cell(row=row, column=col))

                if class_name:
                    effective_time = class_time if class_time else last_known_time
                    classes.append((current_day, effective_time, class_name))

    return classes


def parse_time(time_str):
    time_str = re.sub(r'^[IVXLCDM]+\s*\n?', '', time_str).strip()

    time_str = re.sub(r'[-–]', '-', time_str)

    time_match = re.search(r'(\d{3,4})\s*-\s*(\d{3,4})', time_str)
    if time_match:
        start_time, end_time = time_match.groups()

        if len(start_time) == 3:
            start_time = f"0{start_time[:1]}:{start_time[1:]}"
        else:
            start_time = f"{start_time[:2]}:{start_time[2:]}"

        if len(end_time) == 3:
            end_time = f"0{end_time[:1]}:{end_time[1:]}"
        else:
            end_time = f"{end_time[:2]}:{end_time[2:]}"

        return start_time, end_time

    return None, None


def clean_class_name(class_name):
    return re.sub(r'\s+', ' ', class_name).strip()


def create_ics_file(schedule, output_file="schedule.ics"):
    cal = Calendar()

    for day, time, class_info in schedule:
        start_time, end_time = parse_time(time)
        if not start_time or not end_time:
            print(f"Skipping due to incorrect time format: {time}")
            continue

        class_name = clean_class_name(class_info)

        event_date = datetime(2025, 2, 17)
        weekday_mapping = {
            "Понеділок": 0, "Вівторок": 1, "Середа": 2, "Четвер": 3, "П’ятниця": 4, "Субота": 5, "Неділя": 6
        }
        event_date += timedelta(days=weekday_mapping.get(day, 0))

        print(
            f"Creating event: {class_name}, Date: {event_date.strftime('%Y-%m-%d')}, Start: {start_time}, End: {end_time}")

        event_start_dt = datetime.strptime(f"{event_date.strftime('%Y-%m-%d')} {start_time}", "%Y-%m-%d %H:%M")
        event_end_dt = datetime.strptime(f"{event_date.strftime('%Y-%m-%d')} {end_time}", "%Y-%m-%d %H:%M")

        if event_end_dt <= event_start_dt:
            print(f"Skipping event due to invalid time range: {start_time} - {end_time}")
            continue

        event = Event()
        event.name = class_name
        event.begin = event_start_dt
        event.end = event_end_dt
        event.description = f"Class: {class_name}"
        cal.events.add(event)

    with open(output_file, "w", encoding="utf-8") as f:
        f.writelines(cal)

    print(f"ICS file created: {output_file}")


if __name__ == "__main__":
    wb = openpyxl.load_workbook("schedules/Розклад_2022_2023_І_семестр_Інф_ІІ_курс1.xlsx")
    sheet = wb.active

    group_names = extract_groups(sheet)
    print("Group names:", group_names)
    #
    # days = extract_days(sheet)
    # print("Days:", days)
    #
    # class_times = extract_class_times(sheet)
    # print("Class times:", class_times)
    #
    # day = "Четвер"
    # class_times = extract_class_times_by_day(sheet, day)
    # print(f"Class times for {day}: {class_times}")
    #
    # group = "ПМА – 21с"
    # classes = extract_classes_by_group(sheet, group)
    # print(f"Classes for {group}:")
    # for study_class in classes:
    #     print(study_class)

    group_name = "ПМА – 21с"
    schedule = extract_classes_by_group(sheet, group_name)

    create_ics_file(schedule)
